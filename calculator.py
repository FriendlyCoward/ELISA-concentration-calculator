import pandas as pd
import numpy as np
import re
import matplotlib.pyplot as plt
import os
import scipy.optimize
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

def control_graph(parameters: (float), graph_name: str, concentration_data, intensity_data):
    dummy_data = np.linspace(np.floor(np.min(concentration_data)), np.ceil(np.max(concentration_data)), 5000)
    dummy_model = sigmoid(parameters, dummy_data)
    plt.plot(concentration_data, intensity_data, marker = ".", linestyle = "None")
    plt.plot(dummy_data, dummy_model)
    plt.title("Plate well colour intensity based on analyte concentration")
    plt.ylabel("Colour intensity, A")
    plt.xlabel("Concentration, log10(ug/ml)")
    plt.grid()
    plt.savefig(f"./pictures/{graph_name}.png")
    plt.close()

# sigmoid() is for logistical regression. 
def sigmoid(parameters: [float], xdata) -> float:
    x0, y0, c, k = parameters
    y = c / (1 + np.exp(-k * (xdata - x0))) + y0
    return y

# sigmoid_reverse() is for predicting unknown concentrations from color intensity.
def sigmoid_reverse(parameters: [float], ydata) -> float:
    x0, y0, c, k = parameters
    x = x0 - ((np.log((c / (ydata - y0)) - 1)) / k)
    return x

# loss() is used only for fitting sigmoid()
def loss(parameters: [float], xdata, ydata) -> float:
    return ydata - sigmoid(parameters, xdata)

pattern = r".*xlsx$"
data_files = [file for file in os.listdir("./data") 
              if (os.path.isfile(os.path.join("./data", file)) and re.match(pattern, file) != None)]
plate_dictionary = {}

for file in data_files: 
    file_xlsx = pd.ExcelFile(f"./data/{file}")
    pattern = r"Basic Calculation \w*"
    found_nothing = True
    for sheet in file_xlsx.sheet_names:
        match_list = re.match(pattern, sheet)
        if (match_list != None):
            required_sheet = sheet
            found_nothing = False
            break
    if (found_nothing == True):
        continue
    raw_data = pd.read_excel(f"./data/{file}", sheet_name = required_sheet, header = None, index_col = None)

    file_name = list(file)
    file_name = file_name[:-5]
    file_name = ''.join(file_name)
    plate_dictionary[file_name] = {}

    number_of_plates = 1
    indexes = raw_data.index
    pattern = r"Plate"
    for index in indexes:
        cell_text = str(raw_data.loc[index, raw_data.columns[0]])
        match_list = re.findall(pattern, cell_text)
        if (match_list == []):
            continue
        
        plate_name = "plate_" + str(number_of_plates)
        plate_A_row = index + 3
        plate_H_row = index + 11
        plate_data = raw_data.iloc[plate_A_row : plate_H_row]
        plate_data = plate_data.drop(columns = plate_data.columns[0])
        plate_data = plate_data.to_numpy(copy = True, dtype = "float64")
        plate_dictionary[file_name][plate_name] = plate_data
        number_of_plates += 1

template = pd.read_excel("./template.xlsx", sheet_name = "template", header = None, index_col = None, nrows = 8)
template = template.drop(columns = [12])
concentrations = pd.read_excel("./template.xlsx", sheet_name = "concentrations", header = None, index_col = None, nrows = 8)
concentrations = concentrations.drop(columns = [12])

template_experimental = np.where(template == "E", True, False)
template_control = np.where(template == "C", True, False)
concentrations = concentrations.to_numpy(copy = True)

result_dictionary = {}
for file in plate_dictionary.keys():
    result_dictionary[file] = {}
    
for file in plate_dictionary.keys():
    for plate in plate_dictionary[file].keys():
        control_wells = pd.DataFrame(columns = ["concentration", "intensity"])
        for row in range(8):
            for column in range(12):
                if (template_control[row, column] == True):
                    control_wells.loc[len(control_wells), control_wells.columns] = concentrations[row, column], plate_dictionary[file][plate][row, column]
        concentration_data = control_wells["concentration"].to_numpy(copy = True, dtype = "float64")
        concentration_data = np.log10(concentration_data)
        intensity_data = control_wells["intensity"].to_numpy(copy = True, dtype = "float64")
    
        parameter_guess = (np.median(concentration_data), np.median(intensity_data), 1.0, 1.0)
        parameters, ier = scipy.optimize.leastsq(loss, parameter_guess, args = (concentration_data, intensity_data))

        graph_name = f"{file} {plate}"
        control_graph(parameters, graph_name, concentration_data, intensity_data)
        
        result_name = plate + "_results"
        dummy_array = np.zeros((8, 12))
        for row in range(8):
            for column in range(12):
                if (template_experimental[row, column] == True):
                    predicted = sigmoid_reverse(parameters, plate_dictionary[file][plate][row, column])
                    predicted = np.power(10, predicted)
                    dummy_array[row, column] = predicted
                else:
                    dummy_array[row, column] = concentrations[row, column]
        result_dictionary[file][result_name] = dummy_array

for file in result_dictionary.keys():
    for plate in result_dictionary[file].keys():
        result_dictionary[file][plate] = np.round(result_dictionary[file][plate], decimals = 3)

# Make the file
wb = Workbook()
for file in result_dictionary.keys():
    wb.create_sheet(file)
wb.remove(wb["Sheet"])
wb.save("./results.xlsx")
wb.close()

# Write results
with pd.ExcelWriter("./results.xlsx", engine = "openpyxl", mode = "a", if_sheet_exists = "overlay") as writer:
    for file in result_dictionary.keys():
        max_row = -2
        for plate in result_dictionary[file].keys():
            name_df = pd.DataFrame(data = plate, index = ["0"], columns = ["0"])
            data_df = pd.DataFrame(data = result_dictionary[file][plate], index = ["A", "B", "C", "D", "E", "F", "G", "H"], columns = ["1", "2", "3", "4", "5", "6", "7", "8", "9", "10", "11", "12"])
            name_df.to_excel(writer, sheet_name = file, startrow = (max_row + 2 if (max_row != -2) else 0), header = False, index = False)
            max_row += 2
            data_df.to_excel(writer, sheet_name = file, startrow = max_row + 2)
            max_row += 10

# Check for extrapolation and colour the cells
wb = load_workbook("./results.xlsx")
max_concentration = np.round(np.max(control_wells["concentration"]), decimals = 3)
min_concentration = np.round(np.min(control_wells["concentration"]), decimals = 3)
for sheet, file in zip(wb, result_dictionary.keys()):
    current_working_row = -8
    for plate in result_dictionary[file].keys():
        current_row = 0
        current_column = 0
        for row in sheet.iter_rows(min_row = current_working_row + 12,
                                   max_row = current_working_row + 19,
                                   min_col = 2,
                                   max_col = 13):
            for cell in row:
                cell_value = cell.internal_value
                if (cell_value == None and template_experimental[current_row, current_column] == False):
                    continue
                if (cell_value == None and template_experimental[current_row, current_column] == True):
                    cell.fill = PatternFill(start_color = "FF0000", end_color = "FF0000", fill_type = "solid")
                if (cell_value == 0):
                    cell.fill = PatternFill(start_color = "FFFF00", end_color = "FFFF00", fill_type = "solid")
                if (template_control[current_row, current_column] == True):
                    cell.fill = PatternFill(start_color = "90EE90", end_color = "90EE90", fill_type = "solid")
                current_column += 1
            current_row += 1
            current_column = 0
        current_working_row += 12
wb.save("./results.xlsx")
wb.close()

# Writing to .ods
with pd.ExcelWriter("./results.ods", engine = "odf", mode = "w") as writer:
    for file in result_dictionary.keys():
        carrier_df = pd.read_excel("./results.xlsx", sheet_name = f"{file}", header = None, index_col = None)
        carrier_df.to_excel(writer, sheet_name = file, header = None, index = None)